import streamlit as st
from openpyxl import load_workbook
import xlwt
import io
import zipfile
import re
import datetime

st.title("📊 Aplikasi Gabung Data Excel Harga IPH")

# Pilih tahun & bulan
tahun = st.selectbox("Pilih Tahun", [2023, 2024, 2025], index=2, key="tahun")
bulan = st.selectbox(
    "Pilih Bulan",
    ["Januari", "Februari", "Maret", "April", "Mei", "Juni",
     "Juli", "Agustus", "September", "Oktober", "November", "Desember"],
    index=0,
    key="bulan"
)
map_bulan = {
    "Januari": "01", "Februari": "02", "Maret": "03", "April": "04",
    "Mei": "05", "Juni": "06", "Juli": "07", "Agustus": "08",
    "September": "09", "Oktober": "10", "November": "11", "Desember": "12",
}
bulan_num = map_bulan[bulan]

uploaded_files = st.file_uploader(
    "Upload file Excel (.xlsx)",
    type="xlsx",
    accept_multiple_files=True
)

kolom_output = [
    "id", "tahun", "bulan", "minggu",
    "kode_prov", "prov", "nilai_iph", "komoditas",
    "fluktuasi_harga_tertinggi", "disparitas_harga_antar_wilayah", "date_created"
]

# Fungsi ambil minggu dari nama file
def extract_minggu(filename):
    match = re.search(r'M(\d+)', filename)
    if match:
        return int(match.group(1))
    return 1

if st.button("🔄 Proses & Unduh ZIP") and uploaded_files:
    zip_buffer = io.BytesIO()
    with zipfile.ZipFile(zip_buffer, "w") as zf:
        for f in uploaded_files:
            minggu = extract_minggu(f.name)

            wb = load_workbook(f, data_only=True)
            sheetnames = wb.sheetnames

            sheet_kab = wb["360 KabKota"] if "360 KabKota" in sheetnames else None
            sheet_prov = wb["Provinsi"] if "Provinsi" in sheetnames else None

            # ====== STYLE TIAP KOLOM ======
            header_styles = [
                xlwt.easyxf('pattern: pattern solid, fore_colour ocean_blue; font: bold on, colour white;'),
                xlwt.easyxf('pattern: pattern solid, fore_colour light_orange; font: bold on, colour white;'),
                xlwt.easyxf('pattern: pattern solid, fore_colour gold; font: bold on, colour black;'),
                xlwt.easyxf('pattern: pattern solid, fore_colour light_green; font: bold on, colour black;'),
                xlwt.easyxf('pattern: pattern solid, fore_colour rose; font: bold on, colour black;'),
                xlwt.easyxf('pattern: pattern solid, fore_colour turquoise; font: bold on, colour black;'),
                xlwt.easyxf('pattern: pattern solid, fore_colour lime; font: bold on, colour black;'),
                xlwt.easyxf('pattern: pattern solid, fore_colour gray25; font: bold on, colour black;'),
                xlwt.easyxf('pattern: pattern solid, fore_colour tan; font: bold on, colour black;'),
                xlwt.easyxf('pattern: pattern solid, fore_colour ice_blue; font: bold on, colour black;'),
                xlwt.easyxf('pattern: pattern solid, fore_colour yellow; font: bold on, colour black;')
            ]

            body_styles = [
                xlwt.easyxf('pattern: pattern solid, fore_colour pale_blue;'),
                xlwt.easyxf('pattern: pattern solid, fore_colour coral;'),
                xlwt.easyxf('pattern: pattern solid, fore_colour gray40;'),
                xlwt.easyxf('pattern: pattern solid, fore_colour light_green;'),
                xlwt.easyxf('pattern: pattern solid, fore_colour pink;'),
                xlwt.easyxf('pattern: pattern solid, fore_colour light_turquoise;'),
                xlwt.easyxf('pattern: pattern solid, fore_colour lime;'),
                xlwt.easyxf('pattern: pattern solid, fore_colour gray25;'),
                xlwt.easyxf('pattern: pattern solid, fore_colour tan;'),
                xlwt.easyxf('pattern: pattern solid, fore_colour ice_blue;'),
                xlwt.easyxf('pattern: pattern solid, fore_colour yellow;')
            ]

            if sheet_kab:
                rows = list(sheet_kab.iter_rows(values_only=True))
                header_kab = [str(cell) for cell in rows[0]]

                data_kab_final = []
                for i, row in enumerate(rows[1:], 1):
                    if any("Row Label" in str(cell) or "Grand Total" in str(cell) for cell in row):
                        continue

                    baris = {
                        "id": i,
                        "tahun": tahun,
                        "bulan": bulan_num,
                        "minggu": minggu,
                        "kode_prov": row[header_kab.index("kode_kab")],
                        "prov": row[header_kab.index("nama_prov")],
                        "nilai_iph": row[header_kab.index("Perubahan IPH")],
                        "komoditas": row[header_kab.index("Komoditas Andil Besar")],
                        "fluktuasi_harga_tertinggi": row[header_kab.index("Fluktuasi Harga Tertinggi Minggu Berjalan")],
                        "disparitas_harga_antar_wilayah": row[header_kab.index("Disparitas Harga antar Daerah")],
                        "date_created": datetime.datetime.now().strftime("%Y-%m-%d")
                    }
                    data_kab_final.append(baris)

                bk = xlwt.Workbook()
                sk = bk.add_sheet("Gabungan_Kabupaten")

                # Hitung panjang max & set lebar kolom
                max_lens = [len(kol) for kol in kolom_output]
                for baris in data_kab_final:
                    for j, kol in enumerate(kolom_output):
                        panjang = len(str(baris.get(kol, "")))
                        if panjang > max_lens[j]:
                            max_lens[j] = panjang

                for j, kol in enumerate(kolom_output):
                    sk.write(0, j, kol, header_styles[j % len(header_styles)])
                    sk.col(j).width = (max_lens[j] + 2) * 256

                for i, baris in enumerate(data_kab_final, 1):
                    for j, kol in enumerate(kolom_output):
                        sk.write(i, j, baris.get(kol, ""), body_styles[j % len(body_styles)])

                buf = io.BytesIO()
                bk.save(buf)
                buf.seek(0)
                zf.writestr(f"kabupaten_{bulan_num}_{tahun}_M{minggu}.xls", buf.read())

            if sheet_prov:
                rows = list(sheet_prov.iter_rows(values_only=True))
                header_prov = [str(cell) for cell in rows[0]]

                data_prov_final = []
                for i, row in enumerate(rows[1:], 1):
                    if any("Row Label" in str(cell) or "Grand Total" in str(cell) for cell in row):
                        continue

                    baris = {
                        "id": i,
                        "tahun": tahun,
                        "bulan": bulan_num,
                        "minggu": minggu,
                        "kode_prov": row[header_prov.index("kode_prov")],
                        "prov": row[header_prov.index("nama_prov")],
                        "nilai_iph": row[header_prov.index("Perubahan IPH")],
                        "komoditas": row[header_prov.index("Komoditas Andil Terbesar")],
                        "fluktuasi_harga_tertinggi": row[header_prov.index("Fluktuasi Harga Tertinggi Minggu Berjalan")],
                        "disparitas_harga_antar_wilayah": "",
                        "date_created": datetime.datetime.now().strftime("%Y-%m-%d")
                    }
                    data_prov_final.append(baris)

                bp = xlwt.Workbook()
                sp = bp.add_sheet("Gabungan_Provinsi")

                max_lens = [len(kol) for kol in kolom_output]
                for baris in data_prov_final:
                    for j, kol in enumerate(kolom_output):
                        panjang = len(str(baris.get(kol, "")))
                        if panjang > max_lens[j]:
                            max_lens[j] = panjang

                for j, kol in enumerate(kolom_output):
                    sp.write(0, j, kol, header_styles[j % len(header_styles)])
                    sp.col(j).width = (max_lens[j] + 2) * 256

                for i, baris in enumerate(data_prov_final, 1):
                    for j, kol in enumerate(kolom_output):
                        sp.write(i, j, baris.get(kol, ""), body_styles[j % len(body_styles)])

                buf = io.BytesIO()
                bp.save(buf)
                buf.seek(0)
                zf.writestr(f"provinsi_{bulan_num}_{tahun}_M{minggu}.xls", buf.read())

    zip_buffer.seek(0)
    st.download_button(
        "📥 Unduh Gabungan IPH",
        data=zip_buffer,
        file_name=f"gabungan_IPH_{bulan_num}_{tahun}.zip",
        mime="application/zip"
    )
